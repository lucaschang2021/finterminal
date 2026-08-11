# -*- coding: utf-8 -*-
"""Excel 公式单元格保留 + 清洗公式注入中和的回归测试。"""
import os
import shutil
import sys

import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import data_chain as dc
import excel_utils
import mcp_server as m


def _make_xlsx(path, rows):
    pd.DataFrame(rows).to_excel(path, index=False)


def test_read_xlsx_preserves_formula_strings(tmp_path):
    p = str(tmp_path / "f.xlsx")
    _make_xlsx(p, {"x": ["=cmd", "正常", "=1+1", "-3"]})
    df = excel_utils.read_xlsx(p)
    vals = df["x"].tolist()
    assert vals[:2] == ["=cmd", "正常"]
    assert "=1+1" in vals and "-3" in vals
    assert not df["x"].isna().any(), "公式单元格不应被读成 NaN"


def _xlsx_with_cached_formula(path, formula, cached):
    """生成一个公式单元格带缓存计算值的 .xlsx（openpyxl 不写缓存值，需注入 XML）。"""
    import re
    import shutil
    import zipfile

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "val"
    ws["A2"] = formula
    wb.save(path)
    tmp = str(path) + ".tmp"
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith("sheet1.xml"):
                text = data.decode("utf-8")
                m = re.search(r"<f>[^<]+</f>", text)
                if m:
                    text = text.replace(m.group(0), m.group(0) + f"<v>{cached}</v>")
                data = text.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp, path)


def test_read_xlsx_prefers_cached_formula_value(tmp_path):
    """公式单元格有缓存计算值时返回计算值；无缓存时才回退公式字符串。"""
    p = str(tmp_path / "cached.xlsx")
    _xlsx_with_cached_formula(p, "=1+1", 2)
    df = excel_utils.read_xlsx(p)
    assert df["val"].tolist() == [2], df["val"].tolist()

    p2 = str(tmp_path / "nocache.xlsx")
    _make_xlsx(p2, {"val": ["=1+1"]})
    df2 = excel_utils.read_xlsx(p2)
    assert df2["val"].tolist() == ["=1+1"]


def test_clean_neutralizes_xlsx_formula_injection(tmp_path):
    p = str(tmp_path / "inj.xlsx")
    _make_xlsx(p, {"x": ["=cmd", "正常", "-3"]})
    out = m.clean(p)
    # 中和说明出现在报告中；且 =cmd 不再原样展示
    assert "中和公式注入" in out
    assert "=cmd" not in out.replace("'=cmd", "")


def test_load_data_preserves_formula(tmp_path):
    p = str(tmp_path / "d.xlsx")
    _make_xlsx(p, {"x": ["=1+1", "2"]})
    df = m._load_data(p)
    assert df["x"].tolist()[:1] == ["=1+1"]


def test_data_chain_diff_sees_formula_change(tmp_path):
    shutil.rmtree(dc.CHAIN_DIR, ignore_errors=True)
    p = str(tmp_path / "chain.xlsx")
    _make_xlsx(p, {"v": ["=1+1", "2"]})
    dc.record_if_changed(p)
    _make_xlsx(p, {"v": ["=1+2", "2"]})
    r2 = dc.record_if_changed(p)
    assert r2 and r2["action"] == "modified"
    summary = r2["diff"]["summary"]
    assert "修改" in summary or "新增" in summary, summary
    shutil.rmtree(dc.CHAIN_DIR, ignore_errors=True)


def test_knowledge_extract_includes_formula(tmp_path):
    import knowledge as k
    p = str(tmp_path / "kb.xlsx")
    _make_xlsx(p, {"v": ["=1+1"]})
    text = k._extract_text(p)
    assert "=1+1" in text
