"""Excel 读取工具：公式单元格"值优先、无缓存回退公式字符串"。

pandas.read_excel 会把以 = 开头的公式单元格读成 NaN（公式无缓存值时），
导致数据清洗/分析/数据链对比时静默丢失内容。
本模块用 openpyxl 同时读取缓存计算值与公式字符串：
- 公式单元格有缓存计算值（Excel/WPS 保存时通常附带）→ 返回计算值；
- 无缓存值（如程序生成的公式文件）→ 返回公式字符串，既不丢数据也不丢失信息。
"""


def _dedup(cols):
    """重复表头按 pandas 惯例加 .N 后缀（x, x.1, x.2）。"""
    seen = {}
    out = []
    for c in cols:
        if c in seen:
            seen[c] += 1
            out.append(f"{c}.{seen[c]}")
        else:
            seen[c] = 0
            out.append(c)
    return out


def _cached_formula_cells(path, sheet):
    """Return set of cell coordinates whose formula has a cached <v> value.

    openpyxl cannot distinguish "cached value 0" from "no cached value"
    (formulas without cache are read as 0), so we parse the sheet XML:
    only <f> followed by <v> counts as a real cached value.
    """
    import re
    import zipfile

    cached = set()
    try:
        with zipfile.ZipFile(path) as z:
            xml_name = None
            if isinstance(sheet, int):
                xml_name = "xl/worksheets/sheet" + str(sheet + 1) + ".xml"
            else:
                wb_xml = z.read("xl/workbook.xml").decode("utf-8", errors="replace")
                rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="replace")
                rid_map = dict(re.findall(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', wb_xml))
                rid = rid_map.get(sheet)
                if rid:
                    m = re.search(r'Id="' + re.escape(rid) + r'"[^>]*Target="worksheets/([^"]+)"', rels)
                    if m:
                        xml_name = "xl/worksheets/" + m.group(1)
            if not xml_name:
                return cached
            xml = z.read(xml_name).decode("utf-8", errors="replace")
            for m in re.finditer(r'<c[^>]*r="([A-Z]+\d+)"[^>]*>(?:(?!</c>).)*?</c>', xml, re.S):
                cell = m.group(0)
                v = re.search(r'<f[^>]*>.*?</f>\s*<v>([^<]*)</v>', cell, re.S)
                # openpyxl/pandas 写盘时无缓存公式会带 <v>0</v> 或 <v></v>，视为无缓存
                if v and v.group(1).strip() not in ("", "0"):
                    cached.add(m.group(1))
    except Exception:
        pass
    return cached

def read_xlsx(path, sheet=0):
    """读取 .xlsx：公式单元格优先取缓存计算值，无缓存时保留公式字符串。

    参数:
        path: .xlsx 文件路径（支持解密后的临时文件）。
        sheet: sheet 名或从 0 开始的序号，默认第一个 sheet。
    返回:
        pandas.DataFrame；缺失表头按 pandas 惯例命名为 Unnamed: N。
    """
    import openpyxl
    import pandas as pd

    wb_formulas = openpyxl.load_workbook(path, data_only=False, read_only=True)
    wb_values = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws_f = wb_formulas[sheet] if isinstance(sheet, str) else wb_formulas.worksheets[sheet]
        ws_v = wb_values[sheet] if isinstance(sheet, str) else wb_values.worksheets[sheet]
        from itertools import zip_longest

        cached = _cached_formula_cells(path, sheet)
        rows = []
        for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows(values_only=True)):
            cells = []
            for cf, vv in zip_longest(row_f, row_v, fillvalue=None):
                vf = cf.value if cf is not None else None
                if isinstance(vf, str) and vf.startswith("="):
                    # 公式单元格：XML 带缓存计算值才用计算值，无缓存保留公式字符串
                    cells.append(vv if cf is not None and cf.coordinate in cached else vf)
                else:
                    cells.append(vf)
            rows.append(cells)
        if not rows:
            return pd.DataFrame()
        header, data = rows[0], rows[1:]
        cols = [str(h) if h is not None else f"Unnamed: {i}" for i, h in enumerate(header)]
    finally:
        wb_formulas.close()
        wb_values.close()
    if len(set(cols)) != len(cols):
        cols = _dedup(cols)
    return pd.DataFrame(data, columns=cols)
